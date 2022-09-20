import openpyxl
import datetime


class Person(object):
    def __init__(self, name, surname, patronymic, sex, birthday, data_death):
        self.name = name
        self.surname = surname
        self.patronymic = patronymic
        self.sex = sex
        self.birthday = birthday
        self.data_death = data_death
        # Метод . просто метод%) ВОЛШЕБНЫЙ

    def get_full_name(self):
        """
        Метод для получения полного имени по которому потом будет производиться поиск человека в базе данных
        :return:
        """
        return f'{self.surname} {self.name} {self.patronymic}'.lower()

    @staticmethod
    def valid_date(data):
        """
        СтатикМетод для перевода даты в формат datetime для дальнейшего вычисления возраста человека
        :param data:
        :return:
        """
        valid_data_list = []
        date = data.split('.')
        for i in date:
            valid_data_list.append(int(i))
        date_time = datetime.date(valid_data_list[2], valid_data_list[1], valid_data_list[0])
        return date_time

    @staticmethod
    def valid_none(word):
        """
        СтатикМетод для проверки ввода на NoneType и замены его на '__' для корректной работы метода __str__
        :param word:
        :return:
        """
        if word is None:
            return "__"
        return word

    def get_age(self):
        """
        Метод для вычисления возраста человека по дате рождения и дате смерти (если есть)
        :return:
        """
        day_today = datetime.date.today()
        if self.data_death is not None:
            age = str((self.valid_date(self.data_death) - self.valid_date(self.birthday)) / 365)
        else:
            age = str((day_today - self.valid_date(self.birthday)) / 365)
        delta = age.split(',')
        year = delta[0].split(' ')
        return year[0]

    def __str__(self):
        # Делаем изумительно красивы и продуманный вывод;) Все в одну строку, но читаемо и понятно
        return '| Имя: {:<10} | Фамилия: {:<10} | Отчество: {:<10} | '\
               'Пол:{:^2} | Возраст: {:^3} | ' \
               'Родил{} {:^10} | Умер: {:^10}|'.format(self.name, self.valid_none(self.surname),
                                                       self.valid_none(self.patronymic), self.sex,
                                                       self.get_age(), 'cя: ' if self.sex in 'М' else 'ась:',
                                                       self.birthday, self.valid_none(self.data_death))



class PersonList(object):
    file_name = None
    # Присваиваем имени файла по дефолту None дабы избежать ошибки при сохранении файла

    def __init__(self):
        self.persons = []
        # тоже, просто Метод инициализации класса

    def __str__(self):
        return f'Список людей: {self.persons}'
        # Метод для вывода списка людей в строку

    def __repr__(self):
        return f'Список людей: {self.persons}'
        # Метод для вывода списка людей в строку

    def add_person(self, person):
        self.persons.append(person)
        # Метод для добавления человека в список

    def save(self, file_name):
        """
        # Метод для сохранения всего нашего файла в формате xlsx с кастомным или уже существующим именем
        :param file_name:
        :return:
        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Люди'
        sheet.append(['Имя', 'Фамилия', 'Отчество', 'Пол', 'Дата рождения', 'Дата смерти'])
        for person in self.persons:
            sheet.append([person.name, person.surname, person.patronymic, person.sex, person.birthday,
                          person.data_death])
        wb.save(f'{file_name}.xlsx')
        wb.close()
        print("Файл сохранился")


    def load(self, file_name):
        """
        Метод для загрузки файла с уже существующим именем
        :param file_name:
        :return:
        """
        wb = openpyxl.load_workbook(f'{file_name}.xlsx')
        sheet = wb['Люди']
        for row in sheet.iter_rows(min_row=2):
            person = Person(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
            self.persons.append(person)
        wb.close()
        self.file_name = file_name
        print('Загружена база данных')


    def get_info(self):
        """
        Метод для вывода информации о людях в консоль, напоменаем.. 'изумительно красиво и продуманно'
        :return:
        """
        for person in self.persons:
            print(person)


    def find_persons(self, search_name):
        """
        Метод для поиска людей по полному имени или его части
        :param search_name:
        :return:
        """
        persons = []
        for person in self.persons:
            if search_name in person.get_full_name():
                persons.append(person)
        return persons


    def find_person(self, surname):
        """
        Метод для поиска людей по фамилии для того же удаления
        :param surname:
        :return:
        """
        for person in self.persons:
            if person.surname == surname:
                return person
        return None


    def delete_person(self, surname):
        """
        Метод для удаления человека из списка по фамилии
        :param surname:
        :return:
        """
        person = self.find_person(surname)
        if person is not None:
            self.persons.remove(person)
            return True
        return False



def get_valid_date(input_date: str):
    """
    Функция для валидации даты под метод вычисления возраста человека
    :param input_date:
    :return:
    """
    return input_date.replace(" ", ".", 2).replace("/", ".", 2).replace("-", ".", 2)

