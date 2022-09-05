# Прочитать сохранённый csv-файл из задания №17 и сохранить данные в excel-файл,
# кроме возраста – столбец с этими данными не нужен.

import csv
import openpyxl

with open('task_02.csv', 'r', encoding='utf-8', newline='') as f:
    file_writer = csv.reader(f)

    wb = openpyxl.Workbook()
    wb.create_sheet('Учебный лист', 0)
    wb.remove(wb['Sheet'])

    sheet = wb['Учебный лист']
    count = 1
    for col_index, row in enumerate(file_writer):
        while count <= len(row):
            cell = sheet.cell(row=1, column=count + 1)
            cell.value = f"Person {count}"
            count += 1

        for row_index, value in enumerate(row):

            if row_index > 0:
                row_index -= 1
            cell = sheet.cell(row=row_index+2, column=col_index+1)
            cell.value = value
    wb.save('task_04.xlsx')
